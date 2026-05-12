"""
Urna Eletrônica CIPA SENAI — v13
Persistência via Supabase (Postgres + Storage).

Variáveis de ambiente obrigatórias no Render:
  SUPABASE_URL          ex: https://xxxxx.supabase.co
  SUPABASE_SECRET_KEY   sb_secret_... (admin, bypassa RLS)
  FLASK_SECRET_KEY      qualquer string longa aleatória (sessões)
"""
import os, json, csv, io, uuid, hashlib
from datetime import datetime, timezone
from flask import Flask, render_template, request, jsonify, session
from supabase import create_client, Client

# ── Configuração ──────────────────────────────────────────────────────────────

SUPABASE_URL = os.environ.get('SUPABASE_URL', '').strip()
SUPABASE_SECRET_KEY = os.environ.get('SUPABASE_SECRET_KEY', '').strip()

if not SUPABASE_URL or not SUPABASE_SECRET_KEY:
    raise RuntimeError(
        'Variáveis SUPABASE_URL e SUPABASE_SECRET_KEY são obrigatórias. '
        'Configure no Render → Environment.'
    )

sb: Client = create_client(SUPABASE_URL, SUPABASE_SECRET_KEY)
STORAGE_BUCKET = 'candidatos'

app = Flask(__name__)
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'cipa_senai_urna_secret_v13_change_me')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload

DEFAULT_ADMIN_USER = 'admin'
DEFAULT_ADMIN_PASS = 'cipa2025'

# ── Utilitários ───────────────────────────────────────────────────────────────

def hash_senha(senha: str) -> str:
    return hashlib.sha256(senha.encode('utf-8')).hexdigest()

def ts() -> str:
    """Timestamp legível em pt-BR para exibição (não usar como chave)."""
    return datetime.now().strftime('%d/%m/%Y %H:%M:%S')

def now_iso() -> str:
    """Timestamp UTC ISO para o banco."""
    return datetime.now(timezone.utc).isoformat()

def fmt_dt(iso_str):
    """Converte ISO do banco para formato dd/mm/yyyy HH:MM:SS."""
    if not iso_str:
        return ''
    try:
        # Aceita tanto com Z quanto com offset
        s = iso_str.replace('Z', '+00:00') if isinstance(iso_str, str) else iso_str
        dt = datetime.fromisoformat(s) if isinstance(s, str) else s
        # Converte pra horário local (Brasil)
        from datetime import timezone, timedelta
        brt = timezone(timedelta(hours=-3))
        dt_local = dt.astimezone(brt) if dt.tzinfo else dt
        return dt_local.strftime('%d/%m/%Y %H:%M:%S')
    except Exception:
        return str(iso_str)

# ── Mesário system ───────────────────────────────────────────────────────────

def ensure_default_admin():
    """Garante que existe ao menos um admin no banco. Roda na inicialização."""
    try:
        r = sb.table('mesarios').select('usuario').eq('usuario', DEFAULT_ADMIN_USER).execute()
        if not r.data:
            sb.table('mesarios').insert({
                'usuario': DEFAULT_ADMIN_USER,
                'nome': 'Administrador',
                'senha_hash': hash_senha(DEFAULT_ADMIN_PASS),
                'admin': True,
                'senha_temporaria': True,
            }).execute()
            print(f'[init] Admin padrão criado: {DEFAULT_ADMIN_USER}/{DEFAULT_ADMIN_PASS}')
    except Exception as e:
        print(f'[init] Erro ao verificar admin: {e}')

def autenticar_mesario(usuario, senha):
    """Retorna (ok, nome)."""
    usuario = (usuario or '').strip().lower()
    if not usuario:
        return False, None
    r = sb.table('mesarios').select('nome, senha_hash').eq('usuario', usuario).execute()
    if not r.data:
        return False, None
    rec = r.data[0]
    if rec['senha_hash'] == hash_senha(senha):
        return True, rec['nome']
    return False, None

def is_admin(usuario):
    if not usuario:
        return False
    r = sb.table('mesarios').select('admin').eq('usuario', usuario).execute()
    return bool(r.data and r.data[0].get('admin'))

def mesario_logado():
    """Retorna (usuario, nome) ou (None, None)."""
    if 'mesario_usuario' in session and 'mesario_nome' in session:
        return session['mesario_usuario'], session['mesario_nome']
    return None, None

def require_mesario():
    u, _ = mesario_logado()
    return u is not None

# ── CPF global registry ──────────────────────────────────────────────────────

def cpf_has_voted(cpf):
    """Retorna eid se CPF já votou em alguma eleição existente, senão None.
    Também faz limpeza: se a eleição não existe mais, remove o registro."""
    r = sb.table('cpf_registry').select('eleicao_id').eq('cpf', cpf).execute()
    if not r.data:
        return None
    eid = r.data[0]['eleicao_id']
    # Verifica se a eleição ainda existe
    e = sb.table('eleicoes').select('id').eq('id', eid).execute()
    if not e.data:
        # Eleição não existe mais — limpa o registro
        sb.table('cpf_registry').delete().eq('cpf', cpf).execute()
        return None
    return eid

def register_cpf_vote(cpf, eid):
    """Marca o CPF como tendo votado nessa eleição (upsert)."""
    sb.table('cpf_registry').upsert({'cpf': cpf, 'eleicao_id': eid}).execute()

def release_cpfs_for_election(eid):
    """Libera todos os CPFs vinculados a uma eleição (usado em ocultar/excluir)."""
    sb.table('cpf_registry').delete().eq('eleicao_id', eid).execute()

# ── Log helper ───────────────────────────────────────────────────────────────

def append_log(eid, evento, cpf=None):
    """Adiciona evento ao log da eleição. Usa o mesário logado se houver."""
    _, nome = mesario_logado()
    sb.table('logs').insert({
        'eleicao_id': eid,
        'evento': evento,
        'mesario': nome,
        'cpf': cpf,
    }).execute()

# ── Storage helper ───────────────────────────────────────────────────────────

def upload_foto_candidato(eid, file):
    """Faz upload da foto para o bucket 'candidatos' e retorna a URL pública."""
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else 'jpg'
    if ext not in ('jpg', 'jpeg', 'png', 'gif', 'webp'):
        ext = 'jpg'
    fname = f'{eid}/{uuid.uuid4().hex}.{ext}'
    file_bytes = file.read()
    mime_map = {'jpg': 'image/jpeg', 'jpeg': 'image/jpeg', 'png': 'image/png',
                'gif': 'image/gif', 'webp': 'image/webp'}
    try:
        sb.storage.from_(STORAGE_BUCKET).upload(
            path=fname,
            file=file_bytes,
            file_options={'content-type': mime_map.get(ext, 'image/jpeg'),
                          'cache-control': '3600'}
        )
        public_url = sb.storage.from_(STORAGE_BUCKET).get_public_url(fname)
        # Remove ? final que o Supabase às vezes adiciona
        return public_url.rstrip('?')
    except Exception as e:
        print(f'[storage] Erro ao subir foto: {e}')
        return ''

# Inicialização
ensure_default_admin()

# ─── ROUTES: BÁSICO ───────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

# ─── ROUTES: MESÁRIO AUTH ─────────────────────────────────────────────────────

@app.route('/api/mesario/login', methods=['POST'])
def mesario_login():
    dados = request.json or {}
    usuario = (dados.get('usuario') or '').strip().lower()
    senha = dados.get('senha', '')
    ok, nome = autenticar_mesario(usuario, senha)
    if not ok:
        return jsonify({'ok': False, 'msg': 'Usuário ou senha incorretos.'})
    session['mesario_usuario'] = usuario
    session['mesario_nome'] = nome
    session['mesario_admin'] = is_admin(usuario)
    # Senha temporária?
    r = sb.table('mesarios').select('senha_temporaria').eq('usuario', usuario).execute()
    senha_temp = bool(r.data and r.data[0].get('senha_temporaria'))
    return jsonify({
        'ok': True, 'nome': nome, 'admin': session['mesario_admin'],
        'senha_temporaria': senha_temp
    })

@app.route('/api/mesario/logout', methods=['POST'])
def mesario_logout():
    session.pop('mesario_usuario', None)
    session.pop('mesario_nome', None)
    session.pop('mesario_admin', None)
    return jsonify({'ok': True})

@app.route('/api/mesario/me', methods=['GET'])
def mesario_me():
    u, n = mesario_logado()
    if u is None:
        return jsonify({'ok': False})
    r = sb.table('mesarios').select('senha_temporaria').eq('usuario', u).execute()
    senha_temp = bool(r.data and r.data[0].get('senha_temporaria'))
    return jsonify({
        'ok': True, 'usuario': u, 'nome': n,
        'admin': session.get('mesario_admin', False),
        'senha_temporaria': senha_temp
    })

@app.route('/api/mesario/cadastrar', methods=['POST'])
def cadastrar_mesario():
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    dados = request.json or {}
    usuario = (dados.get('usuario') or '').strip().lower()
    nome = (dados.get('nome') or '').strip()
    senha = dados.get('senha', '')
    if not usuario or not nome or not senha:
        return jsonify({'ok': False, 'msg': 'Preencha todos os campos.'})
    if len(senha) < 4:
        return jsonify({'ok': False, 'msg': 'Senha deve ter ao menos 4 caracteres.'})
    if not usuario.replace('_', '').replace('.', '').isalnum():
        return jsonify({'ok': False, 'msg': 'Usuário só pode conter letras, números, _ e .'})

    r = sb.table('mesarios').select('usuario').eq('usuario', usuario).execute()
    if r.data:
        return jsonify({'ok': False, 'msg': 'Este usuário já existe.'})

    sb.table('mesarios').insert({
        'usuario': usuario,
        'nome': nome,
        'senha_hash': hash_senha(senha),
        'admin': False,
        'senha_temporaria': True,
        'criado_por': session.get('mesario_nome', '-'),
    }).execute()
    return jsonify({'ok': True, 'msg': f'Mesário "{nome}" cadastrado. Deverá trocar a senha no primeiro acesso.'})

@app.route('/api/mesario/alterar_senha', methods=['POST'])
def alterar_senha():
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    dados = request.json or {}
    usuario_logado = session.get('mesario_usuario')
    usuario_alvo = (dados.get('usuario') or usuario_logado).strip().lower()
    senha_atual = dados.get('senha_atual', '')
    senha_nova = dados.get('senha_nova', '')
    senha_nova_conf = dados.get('senha_nova_conf', '')

    if not senha_nova or not senha_nova_conf:
        return jsonify({'ok': False, 'msg': 'Informe a nova senha e a confirmação.'})
    if senha_nova != senha_nova_conf:
        return jsonify({'ok': False, 'msg': 'Nova senha e confirmação não coincidem.'})
    if len(senha_nova) < 4:
        return jsonify({'ok': False, 'msg': 'Nova senha deve ter ao menos 4 caracteres.'})

    r = sb.table('mesarios').select('usuario').eq('usuario', usuario_alvo).execute()
    if not r.data:
        return jsonify({'ok': False, 'msg': 'Usuário não encontrado.'})

    is_admin_logado = session.get('mesario_admin', False)
    is_proprio = (usuario_alvo == usuario_logado)

    if not is_proprio and not is_admin_logado:
        return jsonify({'ok': False, 'msg': 'Apenas o administrador pode alterar a senha de outros usuários.'})

    if is_proprio:
        ok, _ = autenticar_mesario(usuario_logado, senha_atual)
        if not ok:
            return jsonify({'ok': False, 'msg': 'Senha atual incorreta.'})
        if senha_nova == senha_atual:
            return jsonify({'ok': False, 'msg': 'A nova senha deve ser diferente da atual.'})

    sb.table('mesarios').update({
        'senha_hash': hash_senha(senha_nova),
        # Se foi o próprio usuário, limpa o flag. Se foi admin redefinindo de outro, força nova troca.
        'senha_temporaria': not is_proprio,
        'senha_alterada_em': now_iso(),
        'senha_alterada_por': session.get('mesario_nome', '-'),
    }).eq('usuario', usuario_alvo).execute()
    return jsonify({'ok': True, 'msg': 'Senha alterada com sucesso.'})

@app.route('/api/mesario/listar', methods=['GET'])
def listar_mesarios():
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    r = sb.table('mesarios').select('usuario, nome, admin, criado_em').execute()
    result = []
    for m in r.data or []:
        result.append({
            'usuario': m['usuario'],
            'nome': m.get('nome', m['usuario']),
            'admin': bool(m.get('admin')),
            'criado_em': fmt_dt(m.get('criado_em')),
        })
    result.sort(key=lambda x: (not x['admin'], x['nome'].upper()))
    return jsonify({'ok': True, 'mesarios': result})

@app.route('/api/mesario/excluir', methods=['POST'])
def excluir_mesario():
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    dados = request.json or {}
    usuario = (dados.get('usuario') or '').strip().lower()
    r = sb.table('mesarios').select('admin').eq('usuario', usuario).execute()
    if not r.data:
        return jsonify({'ok': False, 'msg': 'Mesário não encontrado.'})
    if r.data[0].get('admin'):
        return jsonify({'ok': False, 'msg': 'Não é possível excluir o administrador.'})
    if usuario == session.get('mesario_usuario'):
        return jsonify({'ok': False, 'msg': 'Você não pode excluir a si mesmo.'})
    sb.table('mesarios').delete().eq('usuario', usuario).execute()
    return jsonify({'ok': True})

# ─── ROUTES: ELEIÇÕES (público) ───────────────────────────────────────────────

@app.route('/api/eleicoes', methods=['GET'])
def get_eleicoes():
    """Lista pública: eleições em andamento ou encerradas, não ocultas."""
    r = sb.table('eleicoes').select(
        'id, titulo, eleicao_aberta, eleicao_encerrada, criada_em, encerrada_em'
    ).eq('oculta', False).execute()
    result = []
    for e in r.data or []:
        if e.get('eleicao_aberta') or e.get('eleicao_encerrada'):
            result.append({
                'id': e['id'],
                'titulo': e['titulo'],
                'aberta': e.get('eleicao_aberta', False),
                'encerrada': e.get('eleicao_encerrada', False),
                'criada_em': fmt_dt(e.get('criada_em')),
                'encerrada_em': fmt_dt(e.get('encerrada_em')),
            })
    # Ordena: abertas primeiro, depois encerradas (mais recente primeiro)
    result.sort(key=lambda x: (not x['aberta'], x.get('criada_em', '')), reverse=False)
    return jsonify({'ok': True, 'eleicoes': result})

@app.route('/api/eleicoes/todas', methods=['GET'])
def get_todas_eleicoes():
    """Lista do mesário: todas eleições, incluindo ocultas e pendentes."""
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    r = sb.table('eleicoes').select('*').order('criada_em', desc=True).execute()
    result = []
    for e in r.data or []:
        # Conta total de votos e funcionários
        cands = sb.table('candidatos').select('votos').eq('eleicao_id', e['id']).execute()
        total_cand_votos = sum(c.get('votos', 0) for c in (cands.data or []))
        total_votos = total_cand_votos + (e.get('votos_brancos') or 0)
        fcount = sb.table('funcionarios').select('id', count='exact').eq('eleicao_id', e['id']).execute()
        total_func = fcount.count or 0
        result.append({
            'id': e['id'],
            'titulo': e['titulo'],
            'aberta': e.get('eleicao_aberta', False),
            'encerrada': e.get('eleicao_encerrada', False),
            'oculta': e.get('oculta', False),
            'configurado': e.get('configurado', False),
            'criada_em': fmt_dt(e.get('criada_em')),
            'total_votos': total_votos,
            'tem_planilha': total_func > 0,
        })
    return jsonify({'ok': True, 'eleicoes': result})

@app.route('/api/eleicoes/criar', methods=['POST'])
def criar_eleicao():
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    titulo = (request.form.get('titulo') or '').strip()
    if not titulo:
        return jsonify({'ok': False, 'msg': 'Informe um título.'})

    # Processar planilha (CSV ou XLSX)
    funcionarios = []
    file = request.files.get('planilha')
    if file and file.filename:
        filename = file.filename.lower()
        try:
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(file.read()), read_only=True, data_only=True)
                ws = wb.active
                header_map = {}
                header_row = None
                for row in ws.iter_rows(min_row=1, max_row=10):
                    for cell in row:
                        if cell.value is None: continue
                        k = str(cell.value).strip().upper()
                        if k in ('NOME', 'NOME COMPLETO', 'NOME_COMPLETO', 'FUNCIONARIO', 'FUNCIONÁRIO', 'COLABORADOR'):
                            header_map[cell.column - 1] = 'nome'; header_row = cell.row
                        elif k in ('CPF', 'CPF_FUNCIONARIO', 'CPF_COLABORADOR', 'NR_CPF'):
                            header_map[cell.column - 1] = 'cpf'; header_row = cell.row
                    if header_map: break
                if not header_map or 'nome' not in header_map.values() or 'cpf' not in header_map.values():
                    wb.close()
                    return jsonify({'ok': False, 'msg': 'Colunas NOME e CPF não encontradas na planilha XLSX.'})
                for row in ws.iter_rows(min_row=header_row + 1):
                    nome = cpf = None
                    for cell in row:
                        idx = cell.column - 1
                        if idx not in header_map or cell.value is None: continue
                        v = str(cell.value).strip()
                        if header_map[idx] == 'nome':
                            nome = v if v else None
                        else:
                            cpf_c = v.replace('.', '').replace('-', '').replace(' ', '')
                            if '.' in cpf_c and cpf_c.replace('.', '').isdigit():
                                try: cpf_c = str(int(float(v)))
                                except: pass
                            if cpf_c.isdigit() and len(cpf_c) < 11:
                                cpf_c = cpf_c.zfill(11)
                            cpf = cpf_c
                    if nome and cpf and len(cpf) == 11 and cpf.isdigit():
                        funcionarios.append({'nome': nome, 'cpf': cpf})
                wb.close()
            else:
                raw = file.read()
                for enc in ('utf-8-sig', 'latin-1', 'utf-8'):
                    try:
                        content = raw.decode(enc); break
                    except UnicodeDecodeError:
                        continue
                else:
                    content = raw.decode('utf-8', errors='replace')
                content = content.lstrip('\ufeff')
                first_line = content.split('\n')[0] if content else ''
                if ';' in first_line: delimiter = ';'
                elif '\t' in first_line: delimiter = '\t'
                elif ',' in first_line: delimiter = ','
                else:
                    try: delimiter = csv.Sniffer().sniff(content[:2048]).delimiter
                    except csv.Error: delimiter = ';'
                reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
                for row in reader:
                    nome = cpf = None
                    for key, value in row.items():
                        if key is None or value is None: continue
                        k = key.strip().strip('\ufeff').strip('"').strip("'").strip().upper()
                        v = value.strip().strip('"').strip("'").strip()
                        if k in ('NOME', 'NOME COMPLETO', 'NOME_COMPLETO', 'FUNCIONARIO', 'FUNCIONÁRIO', 'COLABORADOR'):
                            nome = v if v else None
                        elif k in ('CPF', 'CPF_FUNCIONARIO', 'CPF_COLABORADOR', 'NR_CPF'):
                            if v:
                                cpf = v.replace('.', '').replace('-', '').replace(' ', '')
                                if cpf.isdigit() and len(cpf) < 11:
                                    cpf = cpf.zfill(11)
                    if nome and cpf and len(cpf) == 11 and cpf.isdigit():
                        funcionarios.append({'nome': nome, 'cpf': cpf})
                if not funcionarios:
                    return jsonify({'ok': False, 'msg': 'Nenhum funcionário válido encontrado. Verifique se há colunas NOME e CPF.'})
        except Exception as e:
            return jsonify({'ok': False, 'msg': f'Erro ao processar planilha: {e}'})
    else:
        return jsonify({'ok': False, 'msg': 'Anexe a planilha de funcionários (CSV ou XLSX).'})

    # Cria a eleição
    eid = datetime.now().strftime('%Y%m%d_%H%M%S_') + uuid.uuid4().hex[:4]
    sb.table('eleicoes').insert({
        'id': eid, 'titulo': titulo,
        'configurado': False, 'eleicao_aberta': False,
        'eleicao_encerrada': False, 'oculta': False,
        'votos_brancos': 0,
    }).execute()

    # Insere funcionários (em lote)
    func_rows = [{'eleicao_id': eid, 'nome': f['nome'], 'cpf': f['cpf']} for f in funcionarios]
    if func_rows:
        # Insere em chunks de 100 para não estourar limite
        for i in range(0, len(func_rows), 100):
            sb.table('funcionarios').insert(func_rows[i:i+100]).execute()

    append_log(eid, f'Eleição criada: {titulo} — {len(funcionarios)} funcionário(s) cadastrado(s)')
    return jsonify({'ok': True, 'id': eid, 'total_funcionarios': len(funcionarios)})

@app.route('/api/eleicoes/<eid>/status', methods=['GET'])
def status(eid):
    r = sb.table('eleicoes').select('*').eq('id', eid).execute()
    if not r.data:
        return jsonify({'ok': False, 'msg': 'Eleição não encontrada.'})
    e = r.data[0]
    cands = sb.table('candidatos').select('votos').eq('eleicao_id', eid).execute()
    total_cand_votos = sum(c.get('votos', 0) for c in (cands.data or []))
    fcount = sb.table('funcionarios').select('id', count='exact').eq('eleicao_id', eid).execute()
    return jsonify({
        'ok': True, 'titulo': e['titulo'],
        'configurado': e.get('configurado', False),
        'eleicao_aberta': e.get('eleicao_aberta', False),
        'eleicao_encerrada': e.get('eleicao_encerrada', False),
        'oculta': e.get('oculta', False),
        'num_candidatos': len(cands.data or []),
        'total_votos': total_cand_votos + (e.get('votos_brancos') or 0),
        'total_funcionarios': fcount.count or 0,
    })

# ─── ROUTES: MESÁRIO GERENCIAR ELEIÇÕES ───────────────────────────────────────

@app.route('/api/eleicoes/<eid>/configurar', methods=['POST'])
def configurar(eid):
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    r = sb.table('eleicoes').select('*').eq('id', eid).execute()
    if not r.data:
        return jsonify({'ok': False, 'msg': 'Eleição não encontrada.'})
    e = r.data[0]
    if e.get('eleicao_aberta') or e.get('eleicao_encerrada'):
        return jsonify({'ok': False, 'msg': 'Eleição já iniciada ou encerrada.'})

    candidatos_json = request.form.get('candidatos', '[]')
    try:
        candidatos = json.loads(candidatos_json)
    except:
        return jsonify({'ok': False, 'msg': 'Dados de candidatos inválidos.'})

    if len(candidatos) < 1:
        return jsonify({'ok': False, 'msg': 'Informe ao menos 1 candidato.'})
    numeros = [c['numero'] for c in candidatos]
    if len(numeros) != len(set(numeros)):
        return jsonify({'ok': False, 'msg': 'Números de candidatos duplicados.'})

    # Upload fotos e prepara registros
    cand_rows = []
    for i, cand in enumerate(candidatos):
        foto_url = ''
        photo_file = request.files.get(f'foto_{i}')
        if photo_file and photo_file.filename:
            foto_url = upload_foto_candidato(eid, photo_file)
        cand_rows.append({
            'eleicao_id': eid,
            'numero': int(cand['numero']),
            'nome': cand['nome'],
            'foto_url': foto_url,
            'votos': 0,
        })

    # Limpa candidatos antigos (caso reconfigure) e insere os novos
    sb.table('candidatos').delete().eq('eleicao_id', eid).execute()
    sb.table('candidatos').insert(cand_rows).execute()

    sb.table('eleicoes').update({
        'configurado': True,
        'eleicao_aberta': True,
        'votos_brancos': 0,
    }).eq('id', eid).execute()

    # Limpa votos antigos
    sb.table('cpfs_votantes').delete().eq('eleicao_id', eid).execute()

    append_log(eid, f'Eleição iniciada com {len(candidatos)} candidato(s)')
    return jsonify({'ok': True})

@app.route('/api/eleicoes/<eid>/encerrar', methods=['POST'])
def encerrar(eid):
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    r = sb.table('eleicoes').select('id').eq('id', eid).execute()
    if not r.data:
        return jsonify({'ok': False, 'msg': 'Eleição não encontrada.'})
    sb.table('eleicoes').update({
        'eleicao_aberta': False,
        'eleicao_encerrada': True,
        'encerrada_em': now_iso(),
    }).eq('id', eid).execute()
    append_log(eid, 'Eleição encerrada')
    return jsonify({'ok': True})

@app.route('/api/eleicoes/<eid>/ocultar', methods=['POST'])
def ocultar(eid):
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    r = sb.table('eleicoes').select('id').eq('id', eid).execute()
    if not r.data:
        return jsonify({'ok': False, 'msg': 'Eleição não encontrada.'})
    sb.table('eleicoes').update({'oculta': True}).eq('id', eid).execute()
    release_cpfs_for_election(eid)
    append_log(eid, 'Eleição ocultada — CPFs liberados para outras eleições')
    return jsonify({'ok': True})

@app.route('/api/eleicoes/<eid>/exibir', methods=['POST'])
def exibir(eid):
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    r = sb.table('eleicoes').select('id').eq('id', eid).execute()
    if not r.data:
        return jsonify({'ok': False, 'msg': 'Eleição não encontrada.'})
    sb.table('eleicoes').update({'oculta': False}).eq('id', eid).execute()
    append_log(eid, 'Eleição reexibida na tela inicial')
    return jsonify({'ok': True})

@app.route('/api/eleicoes/<eid>/excluir', methods=['POST'])
def excluir_eleicao(eid):
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    dados = request.json or {}
    usuario, nome = mesario_logado()
    ok, _ = autenticar_mesario(usuario, dados.get('senha', ''))
    if not ok:
        return jsonify({'ok': False, 'msg': 'Senha incorreta. A eleição NÃO foi excluída.'})

    r = sb.table('eleicoes').select('*').eq('id', eid).execute()
    if not r.data:
        return jsonify({'ok': False, 'msg': 'Eleição não encontrada.'})
    e = r.data[0]

    # Conta votos e funcionários para o histórico
    cands = sb.table('candidatos').select('votos').eq('eleicao_id', eid).execute()
    total_votos = sum(c.get('votos', 0) for c in (cands.data or [])) + (e.get('votos_brancos') or 0)
    fcount = sb.table('funcionarios').select('id', count='exact').eq('eleicao_id', eid).execute()

    # Arquiva no histórico
    sb.table('historico_eleicoes').insert({
        'id': eid,
        'titulo': e['titulo'],
        'criada_em': e.get('criada_em'),
        'encerrada_em': e.get('encerrada_em'),
        'excluida_por': nome,
        'total_votos': total_votos,
        'total_funcionarios': fcount.count or 0,
        'estava_encerrada': e.get('eleicao_encerrada', False),
    }).execute()

    # Libera CPFs e deleta a eleição (cascata apaga candidatos, funcionários, etc.)
    release_cpfs_for_election(eid)
    sb.table('eleicoes').delete().eq('id', eid).execute()
    # As fotos no Storage ficam órfãs; idealmente um job de limpeza removeria. Por ora, deixamos.

    return jsonify({'ok': True, 'msg': f'Eleição "{e["titulo"]}" excluída.'})

@app.route('/api/historico', methods=['GET'])
def historico_eleicoes():
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    items = []
    # Eleições atuais
    cur = sb.table('eleicoes').select('*').execute()
    for e in cur.data or []:
        cands = sb.table('candidatos').select('votos').eq('eleicao_id', e['id']).execute()
        total = sum(c.get('votos', 0) for c in (cands.data or [])) + (e.get('votos_brancos') or 0)
        fcount = sb.table('funcionarios').select('id', count='exact').eq('eleicao_id', e['id']).execute()
        if e.get('eleicao_encerrada'):
            estado = 'Encerrada'
        elif e.get('eleicao_aberta'):
            estado = 'Em andamento'
        else:
            estado = 'Pendente'
        items.append({
            'id': e['id'],
            'titulo': e['titulo'],
            'estado': estado,
            'criada_em': fmt_dt(e.get('criada_em')),
            'encerrada_em': fmt_dt(e.get('encerrada_em')),
            'total_votos': total,
            'total_funcionarios': fcount.count or 0,
            'oculta': e.get('oculta', False),
            'excluida': False,
        })
    # Excluídas
    h = sb.table('historico_eleicoes').select('*').execute()
    for it in h.data or []:
        items.append({
            'id': it['id'], 'titulo': it['titulo'],
            'estado': 'Excluída',
            'criada_em': fmt_dt(it.get('criada_em')),
            'encerrada_em': fmt_dt(it.get('encerrada_em')),
            'excluida_em': fmt_dt(it.get('excluida_em')),
            'excluida_por': it.get('excluida_por') or '-',
            'total_votos': it.get('total_votos') or 0,
            'total_funcionarios': it.get('total_funcionarios') or 0,
            'oculta': False, 'excluida': True,
        })
    items.sort(key=lambda x: x.get('criada_em', ''), reverse=True)
    return jsonify({'ok': True, 'itens': items})

@app.route('/api/eleicoes/<eid>/zeresima', methods=['GET'])
def zeresima(eid):
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    r = sb.table('eleicoes').select('titulo').eq('id', eid).execute()
    if not r.data:
        return jsonify({'ok': False})
    titulo = r.data[0]['titulo']
    cands = sb.table('candidatos').select('numero, nome').eq('eleicao_id', eid).order('numero').execute()
    append_log(eid, 'Zerésima impressa')
    return jsonify({'ok': True, 'titulo': titulo,
                    'candidatos': cands.data or [], 'gerada_em': ts()})

@app.route('/api/eleicoes/<eid>/log', methods=['GET'])
def get_log(eid):
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    r = sb.table('logs').select('hora, evento, mesario, cpf').eq('eleicao_id', eid).order('hora').execute()
    log = []
    for l in r.data or []:
        log.append({
            'hora': fmt_dt(l.get('hora')),
            'evento': l.get('evento', ''),
            'mesario': l.get('mesario'),
            'cpf': l.get('cpf'),
        })
    return jsonify({'ok': True, 'log': log})

# ─── ROUTES: ELEITOR (votação) ────────────────────────────────────────────────

@app.route('/api/eleicoes/<eid>/validar_cpf', methods=['POST'])
def validar_cpf(eid):
    dados = request.json or {}
    cpf = (dados.get('cpf') or '').replace('.', '').replace('-', '').strip()
    if len(cpf) != 11 or not cpf.isdigit():
        return jsonify({'ok': False, 'msg': 'CPF inválido. Digite 11 dígitos.'})

    r = sb.table('eleicoes').select('eleicao_aberta, titulo').eq('id', eid).execute()
    if not r.data:
        return jsonify({'ok': False, 'msg': 'Eleição não encontrada.'})
    if not r.data[0].get('eleicao_aberta'):
        return jsonify({'ok': False, 'msg': 'Votação não está aberta.'})

    # Verifica se CPF está autorizado nessa eleição
    f = sb.table('funcionarios').select('nome').eq('eleicao_id', eid).eq('cpf', cpf).execute()
    if not f.data:
        return jsonify({'ok': False, 'msg': 'CPF não encontrado na lista de funcionários autorizados.'})

    # Verifica se já votou
    voted_eid = cpf_has_voted(cpf)
    if voted_eid:
        e2 = sb.table('eleicoes').select('titulo').eq('id', voted_eid).execute()
        titulo = e2.data[0]['titulo'] if e2.data else voted_eid
        return jsonify({'ok': False, 'msg': f'Este CPF já votou na eleição "{titulo}".'})

    return jsonify({'ok': True, 'nome_funcionario': f.data[0]['nome']})

@app.route('/api/eleicoes/<eid>/candidato/<numero>', methods=['GET'])
def get_candidato(eid, numero):
    try:
        n = int(numero)
    except ValueError:
        return jsonify({'ok': False})
    c = sb.table('candidatos').select('numero, nome, foto_url').eq('eleicao_id', eid).eq('numero', n).execute()
    if not c.data:
        return jsonify({'ok': False})
    cand = c.data[0]
    return jsonify({'ok': True, 'candidato': {
        'numero': cand['numero'], 'nome': cand['nome'],
        'foto': cand.get('foto_url') or '',
    }})

@app.route('/api/eleicoes/<eid>/candidatos', methods=['GET'])
def get_candidatos_list(eid):
    r = sb.table('eleicoes').select('eleicao_aberta').eq('id', eid).execute()
    if not r.data:
        return jsonify({'ok': False})
    if not r.data[0].get('eleicao_aberta'):
        return jsonify({'ok': False, 'msg': 'Votação não está aberta.'})
    c = sb.table('candidatos').select('numero, nome').eq('eleicao_id', eid).order('numero').execute()
    return jsonify({'ok': True, 'candidatos': c.data or []})

@app.route('/api/eleicoes/<eid>/votar', methods=['POST'])
def votar(eid):
    dados = request.json or {}
    cpf = (dados.get('cpf') or '').replace('.', '').replace('-', '').strip()
    numero = str(dados.get('numero', '')).strip()

    r = sb.table('eleicoes').select('eleicao_aberta, votos_brancos').eq('id', eid).execute()
    if not r.data:
        return jsonify({'ok': False, 'msg': 'Eleição não encontrada.'})
    if not r.data[0].get('eleicao_aberta'):
        return jsonify({'ok': False, 'msg': 'Votação não está aberta.'})

    # CPF autorizado?
    f = sb.table('funcionarios').select('cpf').eq('eleicao_id', eid).eq('cpf', cpf).execute()
    if not f.data:
        return jsonify({'ok': False, 'msg': 'CPF não autorizado.'})

    # Já votou globalmente?
    if cpf_has_voted(cpf):
        return jsonify({'ok': False, 'msg': 'CPF já votou em outra eleição.'})

    cpf_log = cpf[:3] + '.***.***-' + cpf[-2:]

    if numero == '00' or numero == '':
        # Voto em branco — incrementa atomicamente
        atual = r.data[0].get('votos_brancos') or 0
        sb.table('eleicoes').update({'votos_brancos': atual + 1}).eq('id', eid).execute()
        sb.table('cpfs_votantes').insert({'eleicao_id': eid, 'cpf': cpf}).execute()
        register_cpf_vote(cpf, eid)
        append_log(eid, 'Voto em BRANCO registrado', cpf=cpf_log)
        return jsonify({'ok': True, 'branco': True})

    try:
        n = int(numero)
    except ValueError:
        return jsonify({'ok': False, 'msg': 'Número de candidato inválido.'})

    c = sb.table('candidatos').select('id, nome, numero, foto_url, votos').eq('eleicao_id', eid).eq('numero', n).execute()
    if not c.data:
        return jsonify({'ok': False, 'msg': 'Número de candidato inválido.'})
    cand = c.data[0]

    # Incrementa votos do candidato
    sb.table('candidatos').update({'votos': (cand.get('votos') or 0) + 1}).eq('id', cand['id']).execute()
    sb.table('cpfs_votantes').insert({'eleicao_id': eid, 'cpf': cpf}).execute()
    register_cpf_vote(cpf, eid)

    append_log(eid, f'Voto registrado — Nº {n} ({cand["nome"]})', cpf=cpf_log)
    return jsonify({'ok': True, 'branco': False,
                    'candidato': {'numero': cand['numero'], 'nome': cand['nome'],
                                  'foto': cand.get('foto_url') or ''}})

# ─── ROUTES: RESULTADO / RELATÓRIOS ───────────────────────────────────────────

@app.route('/api/eleicoes/<eid>/resultado', methods=['GET'])
def resultado(eid):
    r = sb.table('eleicoes').select('*').eq('id', eid).execute()
    if not r.data:
        return jsonify({'ok': False})
    e = r.data[0]
    # Resultado público só se encerrada; senão exige mesário
    if not e.get('eleicao_encerrada') and not require_mesario():
        return jsonify({'ok': False, 'msg': 'Resultado disponível apenas após encerramento.'})

    cands = sb.table('candidatos').select('numero, nome, foto_url, votos').eq('eleicao_id', eid).execute()
    result = []
    for c in cands.data or []:
        result.append({
            'nome': c['nome'], 'numero': c['numero'],
            'foto': c.get('foto_url') or '',
            'votos': c.get('votos') or 0,
        })
    result.sort(key=lambda x: x['votos'], reverse=True)
    total = sum(x['votos'] for x in result) + (e.get('votos_brancos') or 0)
    fcount = sb.table('funcionarios').select('id', count='exact').eq('eleicao_id', eid).execute()
    return jsonify({
        'ok': True, 'titulo': e['titulo'],
        'candidatos': result,
        'votos_brancos': e.get('votos_brancos') or 0,
        'total_votos': total,
        'total_funcionarios': fcount.count or 0,
        'eleicao_aberta': e.get('eleicao_aberta', False),
        'eleicao_encerrada': e.get('eleicao_encerrada', False),
        'encerrada_em': fmt_dt(e.get('encerrada_em')),
        'criada_em': fmt_dt(e.get('criada_em')),
    })

@app.route('/api/eleicoes/<eid>/relatorio_participacao', methods=['GET'])
def relatorio_participacao(eid):
    if not require_mesario():
        return jsonify({'ok': False, 'msg': 'Acesso restrito.'})
    r = sb.table('eleicoes').select('titulo, criada_em, encerrada_em').eq('id', eid).execute()
    if not r.data:
        return jsonify({'ok': False, 'msg': 'Eleição não encontrada.'})
    e = r.data[0]
    funcs = sb.table('funcionarios').select('nome, cpf').eq('eleicao_id', eid).execute()
    votantes = sb.table('cpfs_votantes').select('cpf').eq('eleicao_id', eid).execute()
    cpfs_votantes = set(v['cpf'] for v in (votantes.data or []))
    votaram, nao_votaram = [], []
    for f in funcs.data or []:
        cpf_mask = f['cpf'][:3] + '.***.***-' + f['cpf'][-2:]
        entry = {'nome': f['nome'], 'cpf_masked': cpf_mask}
        if f['cpf'] in cpfs_votantes:
            votaram.append(entry)
        else:
            nao_votaram.append(entry)
    votaram.sort(key=lambda x: x['nome'].upper())
    nao_votaram.sort(key=lambda x: x['nome'].upper())
    total = len(funcs.data or [])
    return jsonify({
        'ok': True, 'titulo': e['titulo'],
        'votaram': votaram, 'nao_votaram': nao_votaram,
        'total_funcionarios': total,
        'total_votaram': len(votaram),
        'total_nao_votaram': len(nao_votaram),
        'encerrada_em': fmt_dt(e.get('encerrada_em')),
        'criada_em': fmt_dt(e.get('criada_em')),
    })

# ─── ROUTES: TERMO LGPD ───────────────────────────────────────────────────────

@app.route('/api/eleicoes/<eid>/aceitar_termo', methods=['POST'])
def aceitar_termo(eid):
    """Registra a aceitação do termo LGPD pelo eleitor. Auditoria sem expor CPF cru."""
    dados = request.json or {}
    cpf = (dados.get('cpf') or '').replace('.', '').replace('-', '').strip()
    if len(cpf) != 11 or not cpf.isdigit():
        return jsonify({'ok': False, 'msg': 'CPF inválido.'})
    # Hash do CPF para auditoria sem expor o dado pessoal
    cpf_hash = hashlib.sha256(f'{eid}:{cpf}'.encode('utf-8')).hexdigest()
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or '')
    if ',' in ip: ip = ip.split(',')[0].strip()
    ua = request.headers.get('User-Agent', '')[:255]
    try:
        sb.table('termos_aceitos').insert({
            'eleicao_id': eid,
            'cpf_hash': cpf_hash,
            'ip': ip,
            'user_agent': ua,
        }).execute()
    except Exception as e:
        # Falha de auditoria não bloqueia voto (mas registra no console)
        print(f'[lgpd] Erro ao registrar termo: {e}')
    return jsonify({'ok': True})

if __name__ == '__main__':
    app.run(host='0.0.0.0', debug=True, port=5000)
